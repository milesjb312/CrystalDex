#See the README for the functional goals of this program and for author notes and acknowledgements.
#https://byu.app.box.com/developers/console

#Imports
#General imports
import os
#import glob
import shutil
import json
import pandas as pd

import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
#from openpyxl.styles import PatternFill
from datetime import datetime
import time
import threading

#GUI imports
#https://tkdocs.com/tutorial/intro.html#audience
#https://tkdocs.com/tutorial/firstexample.html#design
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog

#Box integration imports:
#https://github.com/box/box-python-sdk-gen/tree/main
#import box_sdk_gen
#from box_sdk_gen import BoxClient, BoxOAuth, OAuthConfig, FileTokenStorage, BoxSDKError, UploadFileAttributes, UploadFileAttributesParentField
import webbrowser

#SebaView integration imports:
#https://codezup.com/automate-windows-tasks-with-python-win32-library/
#https://pywinauto.readthedocs.io/en/latest/getting_started.html
import psutil
import pywinauto
#import pyperclip
from pywinauto import Application
import ctypes

import pyautogui
from pynput import mouse
import pywinauto.keyboard

#Packaging stuff:
#https://realpython.com/pyinstaller-python/

#Paths
script_dir = os.path.dirname(os.path.abspath(__file__))#The directory of this script, so basically the folder where all the code is kept.
server_dir = os.path.join(os.path.abspath("Z:"),"CrystalDex")
os.makedirs(server_dir,exist_ok=True)
crystal_pictures = os.path.join(server_dir,'Resources',"Crystal_Pictures")
os.makedirs(crystal_pictures,exist_ok=True)
server_library = os.path.join(server_dir,"CrystalDex_Library.xlsx")
temp_library = os.path.join(script_dir,"CrystalDex_Library.xlsx")
run_sheet = os.path.join(server_dir,"Run.xlsx")
desktop = os.path.expanduser("~/Desktop")

"""DATABASE MANAGEMENT FUNCTIONS"""

def safe_click(wrapper, coords):
    ctypes.windll.user32.BlockInput(True)
    try:
        time.sleep(0.05)  # let any current input settle
        wrapper.click_input(coords=coords)
        time.sleep(0.05)
    finally:
        ctypes.windll.user32.BlockInput(False)

def connect_to_db():
    return sqlite3.connect(os.path.join(server_dir,"CrystalDex.db"))

def reset_db():
    #This is for resetting the database. The method should not be accessed outside development.
    conn = connect_to_db()#SQLite3 database connection object
    cur = conn.cursor()#Cursor object
    cur.executescript("""
    BEGIN;
    DROP TABLE IF EXISTS crystal_screens;
    DROP TABLE IF EXISTS conditions;
    DROP TABLE IF EXISTS crystal_trays;
    DROP TABLE IF EXISTS crystals;
    COMMIT;
    """)
    conn.commit()

    cur.executescript("""
    BEGIN;
    PRAGMA foreign_keys = ON;
    CREATE TABLE crystal_screens(
                id INTEGER PRIMARY KEY,
                crystal_screen TEXT NOT NULL,
                crystal_screen_symbol TEXT NOT NULL);
    CREATE TABLE conditions(
                id INTEGER PRIMARY KEY,
                crystal_screen_id INTEGER NOT NULL,
                condition_number INTEGER NOT NULL,
                condition TEXT NOT NULL,
                FOREIGN KEY (crystal_screen_id) REFERENCES crystal_screens(id)
                ON DELETE CASCADE);
    CREATE TABLE crystal_trays(
                id INTEGER PRIMARY KEY,
                crystal_screen_id INTEGER NOT NULL,
                date_set TEXT NOT NULL,
                chaperone TEXT,
                crystal_screen TEXT NOT NULL,
                protein TEXT NOT NULL,
                custom_tags TEXT,
                top_left_protein_concentration REAL NOT NULL,
                top_right_protein_concentration REAL NOT NULL,
                bottom_left_protein_concentration REAL NOT NULL);
    CREATE TABLE crystals(
                id INTEGER PRIMARY KEY,
                tray_id INTEGER NOT NULL,
                row TEXT NOT NULL,
                column INT NOT NULL,
                subwell TEXT NOT NULL,
                picture_path TEXT NOT NULL,
                conditions TEXT NOT NULL,
                minor_axis REAL,
                major_axis REAL,
                number_of_crystals INT NOT NULL,
                shape TEXT,
                possible_salt_crystals TEXT NOT NULL,
                precipitation TEXT NOT NULL,
                microcrystals TEXT NOT NULL,
                glassy_protein_or_artifacts TEXT NOT NULL,
                harvester TEXT,
                run INTEGER,
                vial INTEGER,
                port TEXT,
                date_snapped TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (tray_id) REFERENCES crystal_trays(id)
                ON DELETE CASCADE);
    COMMIT;
    """)
    conn.commit()
    conn.close()

#reset_db()

def get_runs():
    conn = connect_to_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    query = f"SELECT DISTINCT run FROM crystals"
    cur.execute(query)
    runs = [int(row[0]) for row in cur.fetchall()]
    conn.close()
    return runs

def get_crystal_screens():
    """Get a list of all crystal screens and their ids."""
    conn = connect_to_db()
    cur = conn.cursor()
    cur.execute("""SELECT id, crystal_screen FROM crystal_screens""")
    crystal_screens = {}
    for r in cur.fetchall():
        crystal_screens[r[1]] = r[0]
    conn.close()
    return crystal_screens

def get_crystal_screen(tray_id):
    conn = connect_to_db()
    cur = conn.cursor()
    cur.execute("""SELECT crystal_screen_id FROM crystal_trays WHERE id = ?""",(tray_id,))
    crystal_screen_id = cur.fetchone()
    conn.close()
    if crystal_screen_id is None:
        return None
    return crystal_screen_id[0]

def get_trays(args_dict=None):
    """Later on, this will act as the filter by python-based filtering of the trays grabbed from the database."""
    conn = connect_to_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    if args_dict!=None:
        clauses = []
        values = []
        for column, value in args_dict.items():
            clauses.append(f"{column} = ?")
            values.append(value)

        query = f"SELECT * FROM crystal_trays"

        query += " WHERE " + " AND ".join(clauses)

        query += " ORDER BY id"

        cur.execute(query, values)
        rows = cur.fetchall()
        conn.close()
    else:
        cur.execute("SELECT * FROM crystal_trays")
        rows = cur.fetchall()
        conn.close()

    trays = [(", ".join(str(item) for item in row)) for row in rows]
    return trays

def get_vials_and_details(run):
    """Runs the following sqlite3 query and returns a dictionary: {"vial":[port,"protein","conditions","notes"]}

    SELECT 
        crystals.vial,
        crystals.port,
        crystal_trays.protein,
        crystals.conditions,
        crystals.notes
        FROM crystals JOIN crystal_trays ON crystals.tray_id = crystal_trays.id WHERE crystals.run=?
    """
    conn = connect_to_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    query = """
    SELECT 
        crystals.vial,
        crystals.port,
        crystal_trays.protein,
        crystals.conditions,
        crystals.notes
        FROM crystals JOIN crystal_trays ON crystals.tray_id = crystal_trays.id WHERE crystals.run=?
    """
    cur.execute(query,(run,))
    vials_and_ports = {row[0]:[row[1],row[2],row[3],row[4]] for row in cur.fetchall()}
    conn.close()
    return vials_and_ports

def get_values(value=None):
    """Later on, this will act as the filter by python-based filtering of the trays grabbed from the database."""
    conn = connect_to_db()
    cur = conn.cursor()
    try:
        if value!=None:
            query = f"SELECT DISTINCT {value} FROM crystal_trays"
            cur.execute(query)
            values = [row[0] for row in cur.fetchall()]
            conn.close()
        return values
    except Exception as e:
        print(f'Search failed. Database may be corrupted.')
        if value!=None:
            query = f"SELECT DISTINCT {value} FROM crystal_trays"
            cur.execute(query)
            values = [row[0] for row in cur.fetchall()]
            conn.close()
        return values

def update_excel():
    """This will update the excel mirror that reflects what's going on in the CrystalDex.db database."""
    conn = connect_to_db()

    df = pd.read_sql_query("""SELECT
        t.id AS tray_id,
        t.crystal_screen_id,
        t.date_set,
        t.chaperone,
        t.crystal_screen,
        t.protein,
        t.custom_tags,
        t.top_left_protein_concentration,
        t.top_right_protein_concentration,
        t.bottom_left_protein_concentration,
        c.id AS crystal_id,
        c.row,
        c.column,
        c.subwell,
        c.picture_path,
        c.conditions,
        c.minor_axis,
        c.major_axis,
        c.number_of_crystals,
        c.shape,
        c.possible_salt_crystals,
        c.precipitation,
        c.microcrystals,
        c.glassy_protein_or_artifacts,
        c.harvester,
        c.run,
        c.vial,
        c.port,
        c.date_snapped,
        c.notes
        FROM crystals c JOIN crystal_trays t ON c.tray_id = t.id
        ORDER BY c.tray_id, c.id""", conn)
    conn.close()

    with pd.ExcelWriter(temp_library, engine="openpyxl",mode="w") as writer:
        if df.empty:
            pd.DataFrame({"info": ["No data available"]}).to_excel(
                writer,
                sheet_name="Empty",
                index=False
            )
        else:
            for tray_id, group in df.groupby("tray_id"):
                # --- Extract tray-level metadata (first row) ---
                row = group.iloc[0]
                group['picture_path'] = group['picture_path'].apply(
                    lambda x: f'=HYPERLINK("{os.path.join(crystal_pictures,x)}", "Open Image")' if pd.notnull(x) else ""
                )

                date_set = row["date_set"]
                protein = str(row["protein"])

                # --- Build sheet name (Excel max 31 chars) ---
                base_name = f"{date_set}_{row["crystal_screen"][0:2]}_{protein}"
                sheet_name = base_name[:31]

                # --- Create header (single-row dataframe) ---
                header_df = pd.DataFrame([{
                    "Tray ID": tray_id,
                    "Date Set": date_set,
                    "Chaperone": row["chaperone"],
                    "Crystal Screen": row["crystal_screen"],
                    "Protein": protein,
                    "Custom Tags": row["custom_tags"],
                    "Top Left Protein Concentration": row['top_left_protein_concentration'],
                    "Top Right Protein Concentration": row['top_right_protein_concentration'],
                    "Bottom Left Protein Concentration": row['bottom_left_protein_concentration']
                }])

                # --- Write to Excel ---
                header_df.to_excel(writer, sheet_name=sheet_name, index=False)

                group.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=3,  # leave space after header
                    index=False
                )
    shutil.move(temp_library,server_library)

"""MICROSCOPE APP FUNCTIONS"""
#In the future, this can be used to allow new users to reconfigure the buttonpresses that are simulated on whatever microscope they're using.
def on_click(x,y,button,pressed):
    global mouse_is_down
    mouse_is_down = pressed

"""MAIN APPLICATION"""

class CrystalDex_main:
    """This is the main class of CrystalDex, containing all the variables that it must pass between different functions, particularly variables that are
    obtained from the user in various tkinter frames or variables involved in window management for the GUI or for the microscope application."""
    def __init__(self):
        self.crystal_size = [0,0]
        self.pixel_to_size = 2000/867 #2 millimeter or 2000 microns per 867 pixels at 100% magnification (ie. a picture size of 1280x960pixels on the screen)

        #Tkinter initializations
        root=tk.Tk()
        self.root = root
        self.root.title("CrystalDex")
        icon_path = os.path.join(script_dir,'Resources',"crystaldex_icon.png")
        icon = tk.PhotoImage(file=icon_path)
        self.root.iconphoto(True,icon)
        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()
        self.root.geometry(f'1050x700+{self.screen_width//2-525}+{self.screen_height//2-350}')

        #Make the window resizable:
        self.root.columnconfigure(0,weight=1)
        self.root.rowconfigure(0,weight=1)
        self.root.protocol("WM_DELETE_WINDOW", self.close_SeBaView_and_root)

        #Track which frame you're in:
        self.current_frame = None
        self.opened_microscope_app = False

        #Window management inits
        self.root.update_idletasks()
        self.root_winfo_id = self.root.winfo_id()
        self.root_wrapper = Application(backend="win32").connect(handle=self.root_winfo_id)
        self.root_window = self.root_wrapper.window(handle=self.root_winfo_id)

    def splash(self):
        self.splash_win = tk.Toplevel(self.root)
        self.splash_win.overrideredirect(True)
        self.splash_win.geometry(f'800x590+{self.screen_width//2-400}+{self.screen_height//2-510//2}')
        splash_path = os.path.join(script_dir,'Resources','CrystalDex_splash.png')
        self.splash_image = tk.PhotoImage(file=splash_path)
        ttk.Label(self.splash_win,text='Loading CrystalDex: DO NOT MOVE THE MOUSE!!!',image=self.splash_image).pack(expand=True)
        self.splash_win.attributes('-topmost',True)
        self.splash_win.lift()
        self.splash_win.focus_set()

    def refocus(self):
        """Refocus the root window if minimized. Whatever frame is currently active will still be visible."""
        #Tkinter internal focusing:
        self.root.deiconify()
        self.root.lift()
        self.root.focus_set()
        pywinauto.keyboard.send_keys('%')

        try:
            if hasattr(self.splash_win,"winfo_exists") and self.splash_win.winfo_exists():
                self.root.after(0,self.splash_win.destroy)
        except Exception:
            pass

    def enforce_position(self):
        self.root.geometry("+0+0")
        self.root.after(100, self.enforce_position)

    def add_menu(self):
        menu = tk.Menu(self.root)
        menu.add_command(label='Home',command=self.startup)
        menu.add_command(label="Help",command=self.Help)
        self.root.config(menu=menu)

    def clear_widgets(self):
        for widget in self.root.winfo_children():
            if isinstance(widget,ttk.Frame):
                widget.destroy()

    def Help(self):
        self.clear_widgets()
        self.add_menu()
        self.root.columnconfigure(0,weight=1)
        self.root.rowconfigure(0,weight=1)
        helpframe = ttk.Frame(self.root,padding='3 3 12 12')
        helpframe.grid(column=0,row=0,sticky='nwes')
        def go_to_docs():
            webbrowser.get('C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe %s').open("https://github.com/milesjb312/CrystalDex")
        ttk.Label(helpframe,text="Welcome to CrystalDex, your helper for recording data from protein crystallization experiments!").grid(column=0,row=0,sticky='new')
        helptext = "This program functions by accessing a server or the cloud and syncing with an sqlite database that contain links to every picture you take." \
        "\nCrystalDex allows you to run the microscope application within its GUI and prompts you to measure and label each crystal."\
        "\nIt then synchronizes all the crystallization screen data from its library of screens with each crystal picture taken."\
        "\nThere are other subprograms in this app that allow you to upload new crystallization screens into its library (such as for optimization screens). "\
        "\nFor more assistance, reach out to miles.j.bradford@outlook.com or take a look at the documentation at: https://github.com/milesjb312/CrystalDex"
        helptext_label = ttk.Label(helpframe,text=helptext)
        helptext_label.grid(column=0,row=1,sticky='new')
        helptext_label.bind("<Button-1>",go_to_docs())
        self.root.after_idle(self.refocus)

    def startup(self):
        self.clear_widgets()
        if not self.opened_microscope_app:
            self.splash()
            threading.Thread(target=self.load_SeBaView,daemon=True).start()
        self.add_menu()
        startup = ttk.Frame(self.root,padding='5 5 20 20')
        self.current_frame = startup
        self.root.geometry(f'{self.screen_width}x{self.screen_height}+0+0')
        self.root.state("zoomed")
        startup.option_add('*tearOFF',tk.FALSE)
        startup.grid(column=0,row=0,sticky='nesw')
        #To make the buttons bigger and prettier, you'll have to use another widget, probably a text widget with a tk.Button placed inside it.
        #https://tkdocs.com/tutorial/text.html#basics
        tk.Button(startup,text="Index New Tray",command=lambda: self.Index_Tray("new"),width=40).grid(column=0,row=0,padx=50,pady=50,sticky='nesw')
        tk.Button(startup,text='Update Tray',command=lambda: self.Index_Tray("old"),width=40).grid(column=1,row=0,padx=50,pady=50,sticky='nesw')
        tk.Button(startup,text='Harvest Crystals',command=lambda: self.Index_Tray("harvesting"),width=40).grid(column=2,row=0,padx=50,pady=50,sticky='nesw')
        tk.Button(startup,text="Upload or Edit Crystal Screen",command=self.Upload_Crystal_Screen,width=40).grid(column=3,row=0,padx=50,pady=50,sticky='nesw')
        tk.Button(startup,text='Design and Upload Custom Screen',command=self.Custom_Screen,width=40).grid(column=0,row=1,padx=50,pady=50,sticky='nesw')
        tk.Button(startup,text='Transfer Crystals',command=self.Crystal_Transfer,width=40).grid(column=1,row=1,padx=50,pady=50,sticky='nesw')

    def load_SeBaView(self):
        """This allows the user to open SeBaView software whenever CrystalDex is running. In the future, I'd like to add a configuration method that lets them choose other
        software and simulate the correct tk.Button presses, but that is currently beyond the scope of this project."""
        exe_path = None
        SeBaView_path = os.path.join(script_dir,'Resources','SeBaView_path_file.json')
        if os.path.exists(SeBaView_path):
            with open(SeBaView_path, "r") as s:
                exe_path = json.load(s).get("SeBaView_path")
        if not exe_path or not os.path.exists(exe_path):
            # Ask user to locate it if not found or invalid
            exe_path = filedialog.askopenfilename(
                title="Select the SeBaView executable",
                filetypes=[("Executable files", "*.exe")]
            )
        if not exe_path:
            print(f'User cancelled the SeBaView path lookup.')
            return
        
        #Save the path.
        with open(SeBaView_path, "w") as s:
            json.dump({"SeBaView_path": exe_path}, s)
        
        exe_name = os.path.basename(exe_path)
        for proc in psutil.process_iter(["name","pid"]):
            try:
                if proc.info["name"] and proc.info["name"].lower() == exe_name.lower():
                    print(f'Closing already running {exe_name} (PID {proc.pid}).')
                    proc.terminate()
                    try:
                        proc.wait(timeout=5)
                        time.sleep(1)
                    except psutil.TimeoutExpired:
                        proc.kill()
                    break
            except (psutil.NoSuchProcess,psutil.AccessDenied):
                continue
        
        self.SeBaView = Application(backend="uia").start(exe_path)
        time.sleep(4)
        try:
            SeBaView_main_window = self.SeBaView.window(title_re=".*SeBaView.*")
            self.SeBaView_wrapper = SeBaView_main_window.wrapper_object()
            self.SeBaView_wrapper.maximize()
            self.SeBaView_wrapper_rect = self.SeBaView_wrapper.rectangle()
            self.SeBaView_wrapper.set_focus()
            safe_click(self.SeBaView_wrapper,(60,165))
            self.SeBaView_wrapper.minimize()
            self.opened_microscope_app = True
        except Exception as e:
            print("Failed to find or focus the SeBaView window. Restarting your computer usually fixes this issue.") #Change this to a Tkinter messagebox
            print(f"Error: {e}")
        self.root.after(0,self.splash_win.destroy)
        self.root.after_idle(self.refocus)

    def close_SeBaView_and_root(self):
        try:
            self.SeBaView_wrapper.close()
        except:
            print(f'Failed to close SeBaView. Do it please!')
        self.root.destroy()

    def monitor_mouse(self):
        """This is just used if you're trying to figure out where you need to simulate a new button click. It is not used in the current program."""
        print(f'monitor_mouse is running...')
        if mouse_is_down:
            x, y = pyautogui.position()
            rel_x = x - self.SeBaView_wrapper_rect.left
            rel_y = y - self.SeBaView_wrapper_rect.top
            print(f'Mouse pressed relative to SeBaView window: ({rel_x}, {rel_y})')
            self.button_location = rel_x,rel_y
        time.sleep(0.1)
        print(f'monitor mouse is done.')

    def add_tray(self,crystal_screen_id,chaperone,crystal_screen,protein,custom_tags,top_left=0,top_right=0,bottom_left=0,date_set="00-00-0000"):
        """This adds a crystal tray to the CrystalDex database. It is usually called by the Index_Tray function and it 
        routes into the characterize_crystals function.
        For reference, this is how the crystal_trays table is made:
        CREATE TABLE crystal_trays(
                id INTEGER PRIMARY KEY,
                crystal_screen_id INTEGER NOT NULL,
                date_set TEXT NOT NULL,
                chaperone TEXT,
                crystal_screen TEXT NOT NULL,
                protein TEXT NOT NULL,
                custom_tags TEXT,
                top_left_protein_concentration REAL NOT NULL,
                top_right_protein_concentration REAL NOT NULL,
                bottom_left_protein_concentration REAL NOT NULL);
        """
        possible_duplicates = get_trays({'date_set':date_set,"crystal_screen":crystal_screen,"top_left_protein_concentration":top_left,"top_right_protein_concentration":top_right,"bottom_left_protein_concentration":bottom_left})
        possible_duplicates = [[possible_duplicate] for possible_duplicate in possible_duplicates]
        if len(possible_duplicates)>0:
            make_new_anyway = messagebox.askyesno(title="Possible Duplicates",message=f"The following tray(s) may be duplicates. Make a new tray anyway? If you choose 'No', you will be routed to the Update Tray function. {possible_duplicates}")
            if not make_new_anyway:
                self.Index_Tray('old')
            else:
                if "" not in [crystal_screen,protein] and sum(1 for conc in [top_left,top_right,bottom_left] if conc!=0)>=1:
                    conn = connect_to_db()
                    cur = conn.cursor()
                    cur.execute("""
                    INSERT INTO crystal_trays
                    (crystal_screen_id, date_set, chaperone, crystal_screen, protein, custom_tags, top_left_protein_concentration, top_right_protein_concentration, bottom_left_protein_concentration)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (crystal_screen_id,date_set,chaperone,crystal_screen,protein,custom_tags,top_left,top_right,bottom_left))
                    self.tray_id = str(cur.lastrowid).strip()
                    conn.commit()
                    conn.close()
                    self.characterize_crystal(method="new")
                else:
                    messagebox.showerror(title="Missing Information",message="Please fill out all required fields and put in at least one protein concentration.")
                    return
        else:
            if "" not in [crystal_screen,protein] and sum(1 for conc in [top_left,top_right,bottom_left] if conc!=0)>=1:
                conn = connect_to_db()
                cur = conn.cursor()
                cur.execute("""
                INSERT INTO crystal_trays
                (crystal_screen_id, date_set, chaperone, crystal_screen, protein, custom_tags, top_left_protein_concentration, top_right_protein_concentration, bottom_left_protein_concentration)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (crystal_screen_id,date_set,chaperone,crystal_screen,protein,custom_tags,top_left,top_right,bottom_left))
                self.tray_id = str(cur.lastrowid).strip()
                conn.commit()
                conn.close()
                self.characterize_crystal(method="new")
            else:
                messagebox.showerror(title="Missing Information",message="Please fill out all required fields and put in at least one protein concentration.")
                return


    def Index_Tray(self,method):
        """This is the GUI method for editing the crystal trays in the database or adding new ones. It is called into from the startup function and routes into the add_tray function or into the edit_tray function."""
        screens = get_crystal_screens()
        if len(screens) == 0:
            messagebox.showerror(title="No Crystal Screens",message="No crystal screens have been uploaded. Upload one and try again.")
            self.startup()
            return
        if method == "new":
            self.clear_widgets()
            self.add_menu()
            new_tray_frame = ttk.Frame(self.root,padding="3 3 12 12")
            self.current_frame = new_tray_frame
            new_tray_frame.grid(column=0,row=0,sticky='nw')

            ttk.Label(new_tray_frame, text="Select from standard tags or type a new entry:").grid(column=1,row=1)

            date_set_values = [str(datetime.now().strftime('%m-%d-%Y'))]
            today_label = ttk.Label(new_tray_frame,text="Today?")
            today_label.grid(column=3,row=5,sticky='nw')
            today_var = tk.BooleanVar()
            today_checkbutton = ttk.Checkbutton(new_tray_frame,variable=today_var,onvalue=True,offvalue=False)
            today_checkbutton.grid(column=4,row=5,sticky='nw')
            date_set_label = ttk.Label(new_tray_frame,text="Date Set (required; 00-00-0000):")
            date_set_label.grid(column=1,row=5,sticky='nw')
            date_set_var = tk.StringVar()
            date_set_drop_down = ttk.Combobox(new_tray_frame,textvariable=date_set_var,values=date_set_values)
            date_set_drop_down.grid(column=2,row=5)
            def set_today(*event):
                date_set_var.set(str(datetime.now().strftime('%m-%d-%Y')))
            today_checkbutton.bind('<ButtonPress>',set_today)

            chaperone_label = ttk.Label(new_tray_frame,text="Crystal Chaperone (optional):")
            chaperone_label.grid(column=1,row=6,sticky='nw')
            chaperone_var = tk.StringVar()
            chaperone_values = get_values('chaperone')
            chaperone_drop_down = ttk.Combobox(new_tray_frame,textvariable=chaperone_var,values=chaperone_values)
            chaperone_drop_down.grid(column=2,row=6)

            crystal_screen_label = ttk.Label(new_tray_frame,text="Crystal Screen (required):")
            crystal_screen_label.grid(column=1,row=7,sticky='nw')
            crystal_screen_var = tk.StringVar()
            crystal_screen_dict = get_crystal_screens()
            crystal_screen_values = list(crystal_screen_dict.keys())
            crystal_screen_drop_down = ttk.Combobox(new_tray_frame,textvariable=crystal_screen_var,values=crystal_screen_values,state="readonly")
            crystal_screen_drop_down.grid(column=2,row=7)

            protein_values = get_values('protein')#Queries the database to find all past proteins used in any crystal tray
            protein_label = ttk.Label(new_tray_frame,text="Target protein: For Moody Lab users, put FULL construct name!!! (do not use special characters /.:;'*?\")")
            protein_label.grid(column=1,row=8,sticky='nw')
            protein_var = tk.StringVar()
            protein_drop_down = ttk.Combobox(new_tray_frame,textvariable=protein_var,values=protein_values)
            protein_drop_down.grid(column=2,row=8,sticky='nw')

            protein_concentration_values = get_values('top_left_protein_concentration')
            protein_concentration_values.extend(get_values('top_right_protein_concentration'))
            protein_concentration_values.extend(get_values('bottom_left_protein_concentration'))
            protein_top_left_concentration_label = ttk.Label(new_tray_frame,text="Target protein stock concentration placed into top left subwell (required):")
            protein_top_left_concentration_label.grid(column=1,row=9,sticky='nw')
            protein_top_left_concentration_var = tk.StringVar()
            protein_top_left_concentration_drop_down = ttk.Combobox(new_tray_frame,textvariable=protein_top_left_concentration_var,values=protein_concentration_values)
            protein_top_left_concentration_drop_down.grid(column=2,row=9,sticky='nw')

            protein_top_right_concentration_label = ttk.Label(new_tray_frame,text="Target protein stock concentration placed into top right subwell (required):")
            protein_top_right_concentration_label.grid(column=1,row=10,sticky='nw')
            protein_top_right_concentration_var = tk.StringVar()
            protein_top_right_concentration_drop_down = ttk.Combobox(new_tray_frame,textvariable=protein_top_right_concentration_var,values=protein_concentration_values)
            protein_top_right_concentration_drop_down.grid(column=2,row=10,sticky='nw')

            protein_bottom_left_concentration_label = ttk.Label(new_tray_frame,text="Target protein stock concentration placed into bottom left subwell (required):")
            protein_bottom_left_concentration_label.grid(column=1,row=11,sticky='nw')
            protein_bottom_left_concentration_var = tk.StringVar()
            protein_bottom_left_concentration_drop_down = ttk.Combobox(new_tray_frame,textvariable=protein_bottom_left_concentration_var,values=protein_concentration_values)
            protein_bottom_left_concentration_drop_down.grid(column=2,row=11,sticky='nw')

            custom_tags_values = []
            custom_tags_label = ttk.Label(new_tray_frame,text="Custom Tags (optional; separated by commas, please!):")
            custom_tags_label.grid(column=1,row=12,sticky='nw')
            custom_tags_var = tk.StringVar()
            custom_tags_drop_down = ttk.Combobox(new_tray_frame,textvariable=custom_tags_var,values=custom_tags_values)
            custom_tags_drop_down.grid(column=2,row=12)
            def verify_screen(screen_var):
                if screen_var.get()=="":
                    messagebox.showerror(title="No Crystal Screen",message="You did not enter a crystal screen. Please fill out all requred information and try again.")
                    return
                else:
                    return screen_var.get()
            def verify_conc(conc_var):
                if conc_var.get() == "":
                    return 0
                else:
                    try:
                        return float(conc_var.get())
                    except Exception as e:
                        messagebox.showerror(title="Invalid concentration entry",message="One or more of your concentration entries is invalid. Please ensure that all concentrations are numbers and try again.")
                        return
            def verify_date(date_set_var):
                date_set = date_set_var.get()
                if date_set == "":
                    messagebox.showerror(title="No Date Entry",message="You did not enter a date. Please do so and try again.")
                    return
                else:
                    try:
                        time.strptime(date_set,'%m-%d-%Y')
                        return date_set
                    except Exception:
                        messagebox.showerror(title="Invalid Date",message="Your date entry is invalid. Please try again with the 00-00-0000 format.")
                        return
            tk.Button(new_tray_frame,text="Make new tray",command=lambda: self.add_tray(crystal_screen_dict[verify_screen(crystal_screen_var)],
                                                                                   chaperone_var.get(),
                                                                                   crystal_screen_var.get(),
                                                                                   protein_var.get(),
                                                                                   custom_tags_var.get(),
                                                                                   top_left=verify_conc(protein_top_left_concentration_var),
                                                                                   top_right=verify_conc(protein_top_right_concentration_var),
                                                                                   bottom_left=verify_conc(protein_bottom_left_concentration_var),
                                                                                   date_set=verify_date(date_set_var)
                                                                                   )).grid(column=1,row=13,sticky='W')

            for child in new_tray_frame.winfo_children():
                child.grid_configure(padx=5,pady=5)
            self.root.after_idle(self.refocus)

        elif method=="old" or method=="harvesting":
            """This will let the user edit old trays by changing adding crystal data.
            For reference, this is how the crystal_trays table is made:
            CREATE TABLE crystal_trays(
                    id INTEGER PRIMARY KEY,
                    crystal_screen_id INTEGER NOT NULL,
                    date_set TEXT NOT NULL,
                    chaperone TEXT,
                    crystal_screen TEXT NOT NULL,
                    protein TEXT NOT NULL,
                    custom_tags TEXT,
                    top_left_protein_concentration REAL NOT NULL,
                    top_right_protein_concentration REAL NOT NULL,
                    bottom_left_protein_concentration REAL NOT NULL);
            """
            self.clear_widgets()
            self.add_menu()
            old_tray_frame = ttk.Frame(self.root,padding="3 3 12 12")
            old_tray_frame.grid(column=0,row=0,sticky='nw')
            old_tray_frame.columnconfigure(0,minsize=450)
            tk.Button(old_tray_frame,text="Filter",command=lambda:reget_trays()).grid(column=1,row=3,sticky='W')

            date_set_values = [str(datetime.now().strftime('%m-%d-%Y'))]
            today_label = ttk.Label(old_tray_frame,text="Today?")
            today_label.grid(column=3,row=5,sticky='nw')
            today_var = tk.BooleanVar()
            today_checkbutton = ttk.Checkbutton(old_tray_frame,variable=today_var,onvalue=True,offvalue=False)
            today_checkbutton.grid(column=4,row=5,sticky='nw')
            date_set_label = ttk.Label(old_tray_frame,text="Date Set (required; 00-00-0000):")
            date_set_label.grid(column=1,row=5,sticky='nw')
            date_set_var = tk.StringVar()
            date_set_drop_down = ttk.Combobox(old_tray_frame,textvariable=date_set_var,values=date_set_values)
            date_set_drop_down.grid(column=2,row=5)
            def set_today(*event):
                date_set_var.set(str(datetime.now().strftime('%m-%d-%Y')))
            today_checkbutton.bind('<ButtonPress>',set_today)

            chaperone_label = ttk.Label(old_tray_frame,text="Crystal Chaperone (optional):")
            chaperone_label.grid(column=1,row=6,sticky='nw')
            chaperone_var = tk.StringVar()
            chaperone_values = get_values('chaperone')
            chaperone_drop_down = ttk.Combobox(old_tray_frame,textvariable=chaperone_var,values=chaperone_values)
            chaperone_drop_down.grid(column=2,row=6)

            crystal_screen_label = ttk.Label(old_tray_frame,text="Crystal Screen (required):")
            crystal_screen_label.grid(column=1,row=7,sticky='nw')
            crystal_screen_var = tk.StringVar()
            crystal_screen_dict = get_crystal_screens()
            crystal_screen_values = list(crystal_screen_dict.keys())
            crystal_screen_drop_down = ttk.Combobox(old_tray_frame,textvariable=crystal_screen_var,values=crystal_screen_values,state="readonly")
            crystal_screen_drop_down.grid(column=2,row=7)

            protein_values = get_values('protein')#Queries the database to find all past proteins used in any crystal tray
            protein_label = ttk.Label(old_tray_frame,text="Target protein: For Moody Lab users, put FULL construct name!!! (do not use special characters /.:;'*?\")")
            protein_label.grid(column=1,row=8,sticky='nw')
            protein_var = tk.StringVar()
            protein_drop_down = ttk.Combobox(old_tray_frame,textvariable=protein_var,values=protein_values)
            protein_drop_down.grid(column=2,row=8,sticky='nw')

            st_name_label = ttk.Label(old_tray_frame,text=('Please select a tray to edit.'))
            st_name_label.grid(column=0,row=4)
            tray_var = tk.StringVar()
            trays = get_trays()
            st_name_combobox = ttk.Combobox(old_tray_frame,textvariable=tray_var,values=trays,state="readonly")
            st_name_combobox.grid(column=0,row=5,sticky='ew')

            if method=="harvesting":
                run_label = ttk.Label(old_tray_frame,text='Beamline Run:')
                run_label.grid(column=0,row=6)
                runs = get_runs()
                if len(runs)>0:
                    ttk.Label(old_tray_frame,text=f'{max(runs)}').grid(column=0,row=7)
                    tk.Button(old_tray_frame,text="Harvest from Selected Tray for Current Beamline Run",command=lambda: select_old_tray(tray_var.get(),max(runs))).grid(column=0,row=8)
                    tk.Button(old_tray_frame,text="Start a new run and Harvest from Selected Tray",command=lambda: select_old_tray(tray_var.get(),max(runs)+1)).grid(column=0,row=9)
                else:
                    ttk.Label(old_tray_frame,text="0").grid(column=0,row=7)
                    tk.Button(old_tray_frame,text="Start a new run and Harvest from Selected Tray",command=lambda: select_old_tray(tray_var.get(),1)).grid(column=0,row=9)
            else:
                tk.Button(old_tray_frame,text="Update Selected Tray",command=lambda: select_old_tray(tray_var.get())).grid(column=0,row=6)

            def select_old_tray(tray_var,run=None):
                self.tray_id = tray_var.split(", ")[0]
                if run is not None:
                    if run=="":
                        messagebox.showerror(title="Empty Beamline Run",message="Please enter a valid Beamline Run number.")
                        return
                    else:
                        self.run = int(run)
                        self.characterize_crystal("harvesting")
                else:
                    self.characterize_crystal("old")

            def reget_trays():
                st_name_label = ttk.Label(old_tray_frame,text=('Please select a tray to edit.'))
                st_name_label.grid(column=0,row=4)
                tray_var = tk.StringVar()
                filter = {}
                if date_set_var.get()!="":
                    filter["date_set"]=date_set_var.get()
                if chaperone_var.get()!="":
                    filter["chaperone"]=chaperone_var.get()
                if crystal_screen_var.get()!="":
                    filter["crystal_screen"]=crystal_screen_var.get()
                if protein_var.get()!="":
                    filter["protein"]=protein_var.get()
                trays = get_trays(filter)
                st_name_combobox = ttk.Combobox(old_tray_frame,textvariable=tray_var,values=trays,state="readonly")
                st_name_combobox.grid(column=0,row=5,sticky='ew')

            if method=="old":
                #If the user is certain that none of the trays that show up are theirs:
                none_of_the_above_label = ttk.Label(old_tray_frame,text="If none of the above match your tray, click 'make new tray':")
                none_of_the_above_label.grid(column=3,row=6)
                tk.Button(old_tray_frame,text="make new tray",command=lambda: self.Index_Tray("new")).grid(column=3,row=7,sticky='W')

    def characterize_crystal(self,method):
        """Creates a new crystal entry in the crystals table of the CrystalDex.db. The crystals table is formatted as follows:
        CREATE TABLE crystals(
                id INTEGER PRIMARY KEY,
                tray_id INTEGER NOT NULL,
                row TEXT NOT NULL,
                column INT NOT NULL,
                subwell TEXT NOT NULL,
                picture_path TEXT NOT NULL,
                conditions TEXT NOT NULL,
                minor_axis REAL NOT NULL,
                major_axis REAL NOT NULL,
                number_of_crystals INT NOT NULL,
                shape TEXT NOT NULL,
                possible_salt_crystals TEXT NOT NULL,
                precipitation TEXT NOT NULL,
                microcrystals TEXT NOT NULL,
                glassy_protein_or_artifacts TEXT NOT NULL,
                harvester TEXT,
                run INTEGER,
                vial INTEGER,
                port TEXT,
                date_snapped TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (tray_id) REFERENCES crystal_trays(id)
                ON DELETE CASCADE);"""
        self.SeBaView_wrapper.maximize()
        self.SeBaView_wrapper.set_focus()
        self.clear_widgets()
        self.add_menu()

        self.root.grid_propagate(False)
        self.root.state('normal')
        self.root.geometry(f"{self.screen_width // 4}x{self.screen_height}+0+0")
        crystal_frame = ttk.Frame(self.root,padding="3 3 12 12")
        crystal_frame.grid(column=0,row=0,sticky='nw')
        self.root.columnconfigure(0,weight=1)
        self.root.rowconfigure(0,weight=1)
        self.root.after_idle(self.refocus)
        self.enforce_position()

        ensure_magnified_label = ttk.Label(crystal_frame,text="MAKE SURE the microscope is fully\nmagnified before taking any pictures.\nALSO ENSURE that the SeBaView\n camera is at 80% magnification.")
        ensure_magnified_label.grid(column=1,row=1)

        well_row = tk.StringVar()
        well_row_label = ttk.Label(crystal_frame,text="Well row:")
        well_row_label.grid(column=1,row=2)
        well_row_values = ['A','B','C','D','E','F','G','H']
        well_row_drop_down = ttk.Combobox(crystal_frame,textvariable=well_row,values=well_row_values,state='readonly')
        well_row_drop_down.grid(column=2,row=2)

        well_column = tk.IntVar()
        well_column_label = ttk.Label(crystal_frame,text="Well column:")
        well_column_label.grid(column=1,row=3)
        well_column_values = ['1','2','3','4','5','6','7','8','9','10','11','12']
        well_column_drop_down = ttk.Combobox(crystal_frame,textvariable=well_column,values=well_column_values,state='readonly')
        well_column_drop_down.grid(column=2,row=3)

        subwell = tk.StringVar()
        
        conn = connect_to_db()
        cur = conn.cursor()
        cur.execute("""SELECT DISTINCT top_left_protein_concentration, top_right_protein_concentration, bottom_left_protein_concentration
                    FROM crystal_trays WHERE id=?""",(self.tray_id,))
        rows = cur.fetchall()
        conn.close()
        concs = {}
        for r in rows:
            concs["top_left"] = r[0]
            concs["top_right"] = r[1]
            concs['bottom_left'] = r[2]
        subwell_values = [conc for conc in concs.keys() if concs[conc]!=0]
        subwell_label = ttk.Label(crystal_frame,text="subwell:")
        subwell_label.grid(column=1,row=4)
        subwell_drop_down = ttk.Combobox(crystal_frame,textvariable=subwell,values=subwell_values,state='readonly')
        subwell_drop_down.grid(column=2,row=4)

        crystal_width = tk.StringVar()
        crystal_width_label = ttk.Label(crystal_frame,text='crystal width:')
        crystal_width_label.grid(column=1,row=5)
        crystal_width_entry = ttk.Entry(crystal_frame,textvariable=crystal_width,state=tk.DISABLED)
        crystal_width_entry.grid(column=2,row=5)
        um_width_label = ttk.Label(crystal_frame,text='um')
        um_width_label.grid(column=3,row=5)

        crystal_height = tk.StringVar()
        crystal_height_label = ttk.Label(crystal_frame,text='crystal height:')
        crystal_height_label.grid(column=1,row=6)
        crystal_height_entry = ttk.Entry(crystal_frame,textvariable=crystal_height,state=tk.DISABLED)
        crystal_height_entry.grid(column=2,row=6)
        um_row_label = ttk.Label(crystal_frame,text='um')
        um_row_label.grid(column=3,row=6)

        number_of_crystals = tk.IntVar()
        number_of_crystals_label = ttk.Label(crystal_frame,text='# of harvestable crystals (optional):')
        number_of_crystals_label.grid(column=1,row=7)

        def validate_int_range(new_value):
            if new_value == "":
                return True
            if new_value.isdigit():
                val = int(new_value)
                return 1 <= val <= 12  # enforce range
            return False
        vcmd = (self.root.register(validate_int_range), "%P")

        number_of_crystals_entry = ttk.Spinbox(crystal_frame,from_=0,to=100,textvariable=number_of_crystals,validate="key",validatecommand=vcmd)
        number_of_crystals_entry.grid(column=2,row=7)

        shape = tk.StringVar()
        shape_label = ttk.Label(crystal_frame,text='Shape of crystals:')
        shape_label.grid(column=1,row=8)
        shape_entry = ttk.Entry(crystal_frame,textvariable=shape)
        shape_entry.grid(column=2,row=8)

        possible_salt_crystals = tk.BooleanVar()
        possible_salt_crystals_label = ttk.Label(crystal_frame,text="Possibly a salt crystal")
        possible_salt_crystals_label.grid(column=1,row=9)
        ttk.Checkbutton(crystal_frame,variable=possible_salt_crystals,onvalue=True,offvalue=False).grid(column=2,row=9)

        precipitation = tk.BooleanVar()
        precipitation_label = ttk.Label(crystal_frame,text="Precipitation present")
        precipitation_label.grid(column=1,row=10)
        ttk.Checkbutton(crystal_frame,variable=precipitation,onvalue=True,offvalue=False).grid(column=2,row=10)

        microcrystals = tk.BooleanVar()
        microcrystals_label = ttk.Label(crystal_frame,text="Microcrystals present")
        microcrystals_label.grid(column=1,row=11)
        ttk.Checkbutton(crystal_frame,variable=microcrystals,onvalue=True,offvalue=False).grid(column=2,row=11)

        glassy_protein_or_artifacts = tk.BooleanVar()
        glassy_protein_or_artifacts_label = ttk.Label(crystal_frame,text="Glassy protein or artifacts present")
        glassy_protein_or_artifacts_label.grid(column=1,row=12)
        ttk.Checkbutton(crystal_frame,variable=glassy_protein_or_artifacts,onvalue=True,offvalue=False).grid(column=2,row=12)

        x = 0
        if method=="harvesting":
            x = 2
            harvester = tk.StringVar()
            harvester_label = ttk.Label(crystal_frame,text='Full name of harvester:')
            harvester_label.grid(column=1,row=13)
            harvester_entry = ttk.Entry(crystal_frame,textvariable=harvester)
            harvester_entry.grid(column=2,row=13)

            vial = tk.StringVar()
            vial_label = ttk.Label(crystal_frame,text='Enter vial number:')
            vial_label.grid(column=1,row=14)
            vials_available=[i for i in range(1,401)]
            vial_dropdown = ttk.Combobox(crystal_frame,textvariable=vial,values=vials_available,state='readonly')
            vial_dropdown.grid(column=2,row=14)

        notes_label = ttk.Label(crystal_frame,text="Crystallographer notes:")
        notes_label.grid(column=1,row=13+x)
        notes = tk.Text(crystal_frame, width = 50, height = 5)
        notes.grid(column=1,row=14+x,columnspan=2)

        def get_condition_number(well_row,well_column):
            if well_row!="" and well_column!=0:
                return (column_index_from_string(well_row)-1)*12+well_column
            else:
                return

        def get_condition():
            print(f'tray_id: {self.tray_id}')
            conn = connect_to_db()
            cur = conn.cursor()
            crystal_screen = get_crystal_screen(self.tray_id)
            condition_number = get_condition_number(well_row.get(),int(well_column.get()))
            print(f'condition_number: {well_row.get(),well_column.get(),condition_number}')
            cur.execute("""SELECT condition FROM conditions WHERE crystal_screen_id = ? AND condition_number = ?""",(crystal_screen,condition_number))
            condition_args = cur.fetchone()
            conn.close()
            if condition_args is None:
                return ""
            return condition_args[0]

        def update_crystal_size_vars():
            crystal_width.set(f'{self.crystal_size[0]}')
            crystal_height.set(f'{self.crystal_size [1]}')

        def get_minor_axis():
            minor_axis = min(crystal_height.get(),crystal_width.get())
            return minor_axis
        
        def get_major_axis():
            major_axis = max(crystal_height.get(),crystal_width.get())
            return major_axis
        
        def get_date_snapped():
            date_snapped = str(datetime.now().strftime('%m-%d-%Y-%H-%M-%S'))
            return date_snapped
        
        def get_harvester():
            if method=="harvesting":
                return harvester.get()
            return "None"

        tk.Button(crystal_frame,text ='Measure Crystal',
                   command=lambda: self.measure_crystal(update_crystal_size_vars)).grid(column=1,row=15+x)
        
        if method=="harvesting":
            tk.Button(crystal_frame,text = 'Take picture and proceed to next well',
                      command = lambda: self.take_picture(well_row.get(),well_column.get(),subwell.get(),get_condition(),get_minor_axis(),get_major_axis(),number_of_crystals.get(),shape.get(),possible_salt_crystals.get(),precipitation.get(),microcrystals.get(),glassy_protein_or_artifacts.get(),get_harvester(),self.run,int(vial.get()),get_date_snapped(),notes.get("1.0", "end-1c"),method)).grid(column=1,row=16+x)
            tk.Button(crystal_frame,text = "Delete previous picture (crystal was not harvested)",
                      command = lambda: self.delete_picture()).grid(column=1,row=17+x)
            x+=1
        else:
            tk.Button(crystal_frame,text = 'Take picture and proceed to next well',
                      command = lambda: self.take_picture(well_row.get(),well_column.get(),subwell.get(),get_condition(),get_minor_axis(),get_major_axis(),number_of_crystals.get(),shape.get(),possible_salt_crystals.get(),precipitation.get(),microcrystals.get(),glassy_protein_or_artifacts.get(),get_harvester(),0,0,get_date_snapped(),notes.get("1.0", "end-1c"),method)).grid(column=1,row=16+x)

        tk.Button(crystal_frame,text="Done with this tray",
                   command=lambda: self.startup()).grid(column=1,row=17+x)
        
        for child in crystal_frame.winfo_children():
            child.grid_configure(padx=5,pady=10)
    
    def delete_picture(self,crystal_id=""):
        crystal_id = crystal_id
        if crystal_id=="":
            if not self.picture_taken:
                return
            else:
                conn = connect_to_db()
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                query = """SELECT * FROM crystals ORDER BY id DESC LIMIT 1"""
                cur.execute(query)
                crystal_id = cur.fetchone()[0]
                conn.close()

        conn = connect_to_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        query = """DELETE FROM crystals where id = ?"""
        print(f'crystal_id to delete: {crystal_id}')
        cur.execute(query,(crystal_id,))
        conn.close()
            
    def take_picture(self,row,column,subwell,condition,minor_axis,major_axis,number_of_crystals,shape,possible_salt_crystals,precipitation,microcrystals,glassy_protein_or_artifacts,harvester,run,vial,date_snapped,notes, method):
        """This is the pride and jewel of CrystalDex, which allows users to take a picture, name it, and upload it all at once without any extra hassle."""
        if "" not in [row,subwell,condition] and column!=0:
            pass
        else:
            messagebox.showerror(title="Incomplete Characterization",message="Please fill out all required fields and try again.")
            return
        if method=="harvesting":
            if self.crystal_size[1] == 0:
                messagebox.showerror(title='No crystal measurement',message="You haven't measured your crystal, silly! Try again.")
                return
        image_title = f'{self.tray_id}_{row}{column}_{subwell}_{date_snapped}'
        self.SeBaView_wrapper.maximize()
        self.SeBaView_wrapper.set_focus()
        safe_click(self.SeBaView_wrapper,(55,70))
        time.sleep(2)
        safe_click(self.SeBaView_wrapper,(750,450))#This is supposed to access the Desktop tk.Button to save the photos.
        time.sleep(1)
        desktop_files = [f for f in os.listdir(desktop) if os.path.isfile(os.path.join(desktop, f))]
        prev_len_desktop_files = len(desktop_files)
        pywinauto.keyboard.send_keys(f"{image_title}{{ENTER}}") #Enter the image_title name into the save window
        unsaved = True
        while unsaved:
            desktop_files = [f for f in os.listdir(desktop) if os.path.isfile(os.path.join(desktop, f))]
            len_desktop_files = len(desktop_files)
            if prev_len_desktop_files != len_desktop_files:
                unsaved=False
                time.sleep(0.1)
        
        suffixes = ['.jpeg','.jpg','.bmp','.tif','']
        file_path_possibilities = [os.path.join(desktop,image_title+suffix) for suffix in suffixes]
        moved = False
        for file_path in file_path_possibilities:
            if os.path.exists(file_path):
                try:
                    picture_path = shutil.move(file_path, crystal_pictures)
                    picture_path = os.path.basename(picture_path)
                    moved = True
                    break
                except Exception as e:
                    messagebox.showerror(title="Failed to Upload a File",message=f"CrystalDex couldn't move the file: {file_path} into its internal resources because of the error: {e}")
        if not moved:
            messagebox.showerror(title="Failed to Locate File",message=f"CrystalDex couldn't find the file for the picture: {image_title}. Did you ensure that it saved to desktop?")

        self.root.after_idle(self.refocus)

        """ The crystals table is formatted as follows:
        CREATE TABLE crystals(
                id INTEGER PRIMARY KEY,
                tray_id INTEGER NOT NULL,
                row TEXT NOT NULL,
                column INT NOT NULL,
                subwell TEXT NOT NULL,
                picture_path TEXT NOT NULL,
                conditions TEXT NOT NULL,
                minor_axis REAL NOT NULL,
                major_axis REAL NOT NULL,
                number_of_crystals INT NOT NULL,
                shape TEXT NOT NULL,
                possible_salt_crystals TEXT NOT NULL,
                precipitation TEXT NOT NULL,
                microcrystals TEXT NOT NULL,
                glassy_protein_or_artifacts TEXT NOT NULL,
                harvester TEXT,
                run INTEGER,
                vial INTEGER,
                port TEXT,
                date_snapped TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (tray_id) REFERENCES crystal_trays(id)
                ON DELETE CASCADE);
        """

        conn = connect_to_db()
        cur = conn.cursor()
        cur.execute("""INSERT INTO crystals (tray_id, row, column, subwell, picture_path, conditions, minor_axis, major_axis, number_of_crystals, shape,
                    possible_salt_crystals, precipitation, microcrystals, glassy_protein_or_artifacts, harvester, run, vial, date_snapped, notes) VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """, (self.tray_id,row,column,subwell,picture_path,condition,minor_axis,major_axis,number_of_crystals,shape,possible_salt_crystals,precipitation,microcrystals,glassy_protein_or_artifacts,harvester,run,vial,date_snapped,notes))
        conn.commit()
        conn.close()
        self.picture_taken = True    

    def measure_crystal(self,function_to_run):
        """This is one of the best features of CrystalDex! However, it does need a calibrate tk.Button. Currently, it only works for the microscope
        in Dr. Moody's lab at BYU.
        """
        self.crystal_size = [0,0]
        if hasattr(self,'measure_tool_window') and self.measure_tool_window.winfo_exists():
            self.measure_tool_window.destroy()
        self.measure_tool_window = tk.Toplevel(self.root)
        icon_path = os.path.join(script_dir,'Resources',"crystaldex_icon.png")
        icon = tk.PhotoImage(file=icon_path)
        self.measure_tool_window.iconphoto(True,icon)
        self.measure_tool_window.title("Crystal Measuring Tool")
        self.measure_tool_window.geometry(f'{self.SeBaView_wrapper_rect.width()-self.screen_width//4-10}x{self.SeBaView_wrapper_rect.height()}+{self.screen_width//4-10}+{0}')
        self.measure_tool_window.resizable(False,False)            
        measure_tool = tk.Canvas(self.measure_tool_window,width=self.measure_tool_window.winfo_width(),height=self.measure_tool_window.winfo_height(),bg='white')
        measure_tool.pack(fill='both',expand=True)
        self.measure_tool_window.attributes('-alpha','0.1')
        self.mouse_pressed = False
        self.line_start = None
        self.current_line = None
        self.measure_tool_window.deiconify()
        self.measure_tool_window.lift()
        self.measure_tool_window.focus_set() #so that when users click and drag, they don't have to click twice on the screen first.
        def on_press(event):
            self.line_start = (event.x,event.y)
        def on_drag(event):
            if self.line_start:
                if self.current_line:
                    measure_tool.delete(self.current_line)
                self.current_line = measure_tool.create_line(
                    self.line_start[0],
                    self.line_start[1],
                    event.x,
                    event.y,
                    fill='blue',
                    width=2
                )
        def on_release(event):
            if self.line_start:
                line_end = (event.x,event.y)
                dx = line_end[0]-self.line_start[0]
                dy = line_end[1]-self.line_start[1]
                length = (dx**2+dy**2)**0.5*self.pixel_to_size
                if self.crystal_size[0] == 0:
                    self.crystal_size[0] = int(length)
                elif self.crystal_size[1] == 0:
                    self.crystal_size[1] = int(length)

                    if hasattr(self, "harvest_crystal_button"):
                        self.harvest_crystal_button.configure(state="normal")
                    if callable(function_to_run):
                        function_to_run()
            self.line_start = None

        measure_tool.bind("<ButtonPress-1>", on_press)
        measure_tool.bind("<B1-Motion>", on_drag)
        measure_tool.bind("<ButtonRelease-1>", on_release)

    def Upload_Crystal_Screen(self):
        """This method allows users to upload a crystal screen directly from Hampton's data sheets."""
        self.clear_widgets()
        self.add_menu()
        self.root.geometry(f'1250x700+{self.screen_width//2-625}+{self.screen_height//2-350}')
        upload_crystal_screen_frame = ttk.Frame(self.root,padding="3 3 12 12")
        upload_crystal_screen_frame.grid(column=0,row=0,sticky='nwes')
        
        crystal_screen_name_label = ttk.Label(upload_crystal_screen_frame,text='Enter the name of the new crystal screen:')
        crystal_screen_name_label.grid(row=0,column=0)
        crystal_screen_name = tk.StringVar()
        crystal_screen_entry = ttk.Entry(upload_crystal_screen_frame,textvariable=crystal_screen_name)
        crystal_screen_entry.grid(row=0,column=1)

        crystal_screen_symbol_label = ttk.Label(upload_crystal_screen_frame,text='Enter 2-letter symbol for new screen:')
        crystal_screen_symbol_label.grid(row=0,column=2)
        crystal_screen_symbol = tk.StringVar()
        crystal_screen_symbol_entry = ttk.Entry(upload_crystal_screen_frame,textvariable=crystal_screen_symbol)
        crystal_screen_symbol_entry.grid(row=0,column=3)

        upload_crystal_screen_button = tk.Button(upload_crystal_screen_frame,text=f"Upload crystal screen",command=lambda: scrape_crystal_screen_data())
        upload_crystal_screen_button.grid(column=4,row=0,sticky='nw')
        upload_crystal_screen_button.configure(text=f'Upload crystal screen')

        def scrape_crystal_screen_data():
            current_screens = get_crystal_screens()
            if crystal_screen_name.get() not in current_screens.keys():
                conn = connect_to_db()
                cur = conn.cursor()
                cur.execute("""INSERT INTO crystal_screens (crystal_screen,crystal_screen_symbol) VALUES (?, ?)""",(crystal_screen_name.get(),crystal_screen_symbol.get()))
                conn.commit()
                conn.close()
            else:
                crystal_screen_id = current_screens[crystal_screen_name.get()]
                conn = connect_to_db()
                cur = conn.cursor()
                print(f'crystal_screen_id: {crystal_screen_id} type: {type(crystal_screen_id)}')
                cur.execute("""DELETE FROM conditions WHERE crystal_screen_id=?""",(str(crystal_screen_id)))
                conn.commit()
                conn.close()

            conditions = {}
            crystal_screen_path = filedialog.askopenfilename(
                    title="Select the excel workbook containing the crystal screen formulation from Hampton. If the workbook contains several pages, you must consolidate them to one page with all the conditions. Delete all rows without a reagent #. Then, delete all extra worksheets."
                )
            
            cswb = openpyxl.load_workbook(crystal_screen_path,read_only=True)
            sheet_names = cswb.sheetnames
            formulation_sheet = cswb[sheet_names[0]]
            print(f'formulation_sheet.title: {formulation_sheet.title}')
            for row in formulation_sheet.iter_rows(min_row=6, max_row=109, min_col=1, max_col=26):
                if row[0].value not in [None, 0]:
                    refs = []
                    for c, cell in enumerate(row, start=1):
                        col_letter = get_column_letter(c)
                        ref = cell.value
                        ref_col = formulation_sheet[f'{col_letter}4'].value
                        if ref:
                            refs.append(str(ref))
                            if ref_col:
                                if ref_col.strip() in ['pH', 'PH', 'Ph', 'ph']:
                                    refs.insert(-1,str(ref_col))
                                elif ref_col.strip() in ['Average','average']:
                                    ref_col_ave = formulation_sheet[f'{col_letter}5'].value
                                    refs.insert(-1,str('Average '+ref_col_ave))

                    condition = " ".join(refs)
                    condition_number = row[0].row-5
                    conditions[condition_number] = condition

            listbox_label = ttk.Label(upload_crystal_screen_frame,text='Review and correct generated conditions:')
            listbox_label.grid(row=2,column=0)
            listbox_values = [f"{condition}" for condition in conditions.values()]
            condition_var = tk.StringVar(value=listbox_values)
            conditions_listbox = tk.Listbox(upload_crystal_screen_frame,listvariable=condition_var,height=25,width=150)
            conditions_listbox.grid(row=3,column=0,columnspan=3)
            
            edited_condition = tk.StringVar()
            condition_entry = tk.Entry(upload_crystal_screen_frame, textvariable=edited_condition, width=150)
            condition_entry.grid(row=4, column=0, columnspan=3)

            self.index = None
            def select_condition(event):
                selection = conditions_listbox.curselection()
                if selection:
                    edited_condition.set(f'{conditions[selection[0]+1]}')
                    self.index = selection[0]

            conditions_listbox.bind('<<ListboxSelect>>',select_condition)

            def overwrite():
                if self.index:
                    text = edited_condition.get()
                    conditions[self.index-1] = text
                    conditions_listbox.delete(self.index)
                    conditions_listbox.insert(self.index, f'{self.index+1} {text}')
                    
            tk.Button(upload_crystal_screen_frame,text='overwrite',command=overwrite).grid(row=4,column=3)
            def save_screens():
                crystal_screens = get_crystal_screens()
                crystal_screen_id = crystal_screens[crystal_screen_name.get()]
                conn = connect_to_db()
                cur = conn.cursor()
                for i, condition in enumerate(conditions.values(),start=1):
                    cur.execute(
                        "INSERT INTO conditions (crystal_screen_id,condition_number,condition) VALUES (?, ?, ?)",
                        (crystal_screen_id, i, condition)
                    )
                conn.commit()
                conn.close()
                self.startup()
                
            tk.Button(upload_crystal_screen_frame,text='Save and finish',command=save_screens).grid(row=5,column=2)

    def Custom_Screen(self):
        """Allows users to either create a custom screen (while optionally looking up a reference condition from any of the screens
        currently in CrystalDex) or to copy information (by hand) from a reference sheet made by Hampton's Make Tray. NOTE: This will not work 
        for any Make Tray optimizations that have conditions that are optimized in a non-linear manner."""
        self.clear_widgets()
        self.add_menu()
        self.root.geometry(f'{self.screen_width}x{self.screen_height}+0+0')
        optimization_screen_frame = ttk.Frame(self.root,padding="3 3 12 12")
        optimization_screen_frame.grid(column=0,row=0,sticky='nwes')
        self.crystal_screen_name = None
        self.crystal_screen_symbol = None
        self.optimization_conditions = {}

        ttk.Label(optimization_screen_frame,text='Fill out the following to name your custom screen. Be advised that CrystalDex appends the date to each custom screen as\n' \
        'this is often one of the most defining characteristics of any tray/screen and helps to avoid duplicate names.',justify='left').grid(column=0,row=0,columnspan=2)

        ttk.Label(optimization_screen_frame,text='Complete name of new custom screen:').grid(row=1,column=0)
        long_name = tk.StringVar()
        long_name_entry = tk.Entry(optimization_screen_frame,textvariable=long_name)
        long_name_entry.grid(row=1,column=1)

        ttk.Label(optimization_screen_frame,text='Two-character code for custom screen:').grid(row=2,column=0)
        two_code = tk.StringVar()
        two_code_entry = tk.Entry(optimization_screen_frame,textvariable=two_code)
        two_code_entry.grid(row=2,column=1)

        tk.Button(optimization_screen_frame,text='Continue',command=lambda: add_screen(long_name_entry.get(),two_code_entry.get())).grid(row=3,column=0)

        def add_screen(crystal_screen_name,crystal_screen_symbol):
            current_screens = get_crystal_screens()
            if crystal_screen_name.get() not in current_screens.keys():
                conn = connect_to_db()
                cur = conn.cursor()
                cur.execute("""INSERT INTO crystal_screens (crystal_screen,crystal_screen_symbol) VALUES (?, ?)""",(crystal_screen_name.get(),crystal_screen_symbol.get()))
                conn.commit()
                conn.close()
            else:
                crystal_screen_id = current_screens[crystal_screen_name.get()]
                conn = connect_to_db()
                cur = conn.cursor()
                print(f'crystal_screen_id: {crystal_screen_id} type: {type(crystal_screen_id)}')
                cur.execute("""DELETE FROM conditions WHERE crystal_screen_id=?""",(str(crystal_screen_id)))
                conn.commit()
                conn.close()

        def select_reference():
            self.clear_widgets()
            self.add_menu()
            self.root.geometry(f'{self.screen_width}x{self.screen_height}+0+0')
            optimization_screen_frame = ttk.Frame(self.root,padding="3 3 12 12")
            optimization_screen_frame.grid(column=0,row=0,sticky='nwes')

            crystal_screens = get_crystal_screens()
            crystal_screens_list = [screen for screen in get_crystal_screens()]
            crystal_screen = tk.StringVar()
            crystal_screens_label = ttk.Label(optimization_screen_frame,text='Available screens:')
            crystal_screens_label.grid(column=0,row=1,sticky='nwes')
            crystal_screens_combobox = ttk.Combobox(optimization_screen_frame,values=crystal_screens_list,textvariable=crystal_screen,height=5,width=100)
            crystal_screens_combobox.grid(column=0,row=2,sticky='nwes')

            ttk.Button(optimization_screen_frame,text="Choose Selected Screen",command=lambda: get_conditions(crystal_screen.get())).grid(column=0,row=3,sticky='nwes')

            def get_conditions(crystal_screen):
                crystal_screen_id = crystal_screens[crystal_screen]
                conn = connect_to_db()
                cur = conn.cursor()
                cur.execute("SELECT condition FROM conditions WHERE crystal_screen_id=?",(crystal_screen_id,))
                lookup_conditions = [str(row[0]) for row in cur.fetchall()]
                conn.close()
                selected_condition_var = tk.StringVar()
                selected_condition = tk.Entry(optimization_screen_frame,textvariable=selected_condition_var)
                selected_condition.grid(column=0,row=5,sticky='nwes')
                lookup_conditions_var = tk.StringVar(value=lookup_conditions)
                lookup_listbox = tk.Listbox(optimization_screen_frame,listvariable=lookup_conditions_var,height=20,width=100)
                lookup_listbox.grid(column=0,row=4,sticky='nwes')

                def select_condition_to_optimize(event):
                    selection = lookup_listbox.curselection()
                    if selection!=-1:
                        index = selection[0]
                        selected_condition_var.set(f'{lookup_conditions[index]}')

                lookup_listbox.bind('<<ListboxSelect>>',select_condition_to_optimize)

                tk.Button(optimization_screen_frame,text='Select condition and continue',command=lambda:optimize(selected_condition=selected_condition_var.get())).grid(row=6,column=0,sticky='nwes')

        def optimize(selected_condition=""):
            self.clear_widgets()
            self.add_menu()
            self.root.geometry(f'{self.screen_width}x{self.screen_height}+0+0')
            optimization_screen_frame = ttk.Frame(self.root,padding="3 3 12 12")
            optimization_screen_frame.grid(column=0,row=0,sticky='nwes')
            reference_label = ttk.Label(optimization_screen_frame,text=f'Reference condition: {selected_condition}')
            reference_label.grid(row=0,column=0,sticky='nwes',columnspan=7)

            tk.Button(optimization_screen_frame,text='Look up new reference',command=lambda: select_reference()).grid(row=1,column=0,sticky='nw')

            ttk.Label(optimization_screen_frame,text="Write a condition and the start, stop, and step concentrations/pH you'd like to iterate that condition over for both the x and y directions. You can populate up to 96 wells. Do not include units in the concentration cells; all units are in molarity or weight percent.").grid(row=6,column=0,columnspan=5,sticky='nw')
            ttk.Label(optimization_screen_frame,text='Please enter in the relevant information for each condition. Ensure that the same number of steps will be generated for your pH and condition settings!').grid(row=7,column=0,columnspan=5,sticky='nwes')

            ttk.Label(optimization_screen_frame,text=f'Steps (optional, up to {96-len(self.optimization_conditions)}, default is 1):').grid(row=9,column=0,sticky='e')
            steps_var = tk.StringVar()
            steps_entry = tk.Entry(optimization_screen_frame,textvariable=steps_var)
            steps_entry.grid(row=9,column=1)

            ttk.Label(optimization_screen_frame,text='Concentration Start (Molar default)').grid(row=9,column=2)
            ttk.Label(optimization_screen_frame,text='Concentration Stop (Molar default)').grid(row=9,column=3)
            ttk.Label(optimization_screen_frame,text='pH Start (None default)').grid(row=9,column=4)
            ttk.Label(optimization_screen_frame,text='pH Stop (None default)').grid(row=9,column=5)

            ttk.Label(optimization_screen_frame,text='Ingredient 0:').grid(row=10,column=0)
            ingredient0_var = tk.StringVar()
            ingredient0_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient0_var)
            ingredient0_entry.grid(row=10,column=1)
            ingredient0_start_var = tk.StringVar()
            ingredient0_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient0_start_var)
            ingredient0_start_entry.grid(row=10,column=2)
            ingredient0_stop_var = tk.StringVar()
            ingredient0_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient0_stop_var)
            ingredient0_stop_entry.grid(row=10,column=3)
            ingredient0_pH_start_var = tk.StringVar()
            ingredient0_pH_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient0_pH_start_var)
            ingredient0_pH_start_entry.grid(row=10,column=4)
            ingredient0_pH_stop_var = tk.StringVar()
            ingredient0_pH_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient0_pH_stop_var)
            ingredient0_pH_stop_entry.grid(row=10,column=5)
            ingredient0_weight_percent_var = tk.BooleanVar(value=False)
            ingredient0_weight_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='weight percent',variable=ingredient0_weight_percent_var,onvalue=True,offvalue=False)
            ingredient0_weight_percent_checkbutton.grid(row=10,column=6)
            ingredient0_volume_percent_var = tk.BooleanVar(value=False)
            ingredient0_volume_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='volume percent',variable=ingredient0_weight_percent_var,onvalue=True,offvalue=False)
            ingredient0_volume_percent_checkbutton.grid(row=10,column=7)

            ttk.Label(optimization_screen_frame,text='Ingredient 1:').grid(row=11,column=0)
            ingredient1_var = tk.StringVar()
            ingredient1_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient1_var)
            ingredient1_entry.grid(row=11,column=1)
            ingredient1_start_var = tk.StringVar()
            ingredient1_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient1_start_var)
            ingredient1_start_entry.grid(row=11,column=2)
            ingredient1_stop_var = tk.StringVar()
            ingredient1_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient1_stop_var)
            ingredient1_stop_entry.grid(row=11,column=3)
            ingredient1_pH_start_var = tk.StringVar()
            ingredient1_pH_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient1_pH_start_var)
            ingredient1_pH_start_entry.grid(row=11,column=4)
            ingredient1_pH_stop_var = tk.StringVar()
            ingredient1_pH_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient1_pH_stop_var)
            ingredient1_pH_stop_entry.grid(row=11,column=5)
            ingredient1_weight_percent_var = tk.BooleanVar(value=False)
            ingredient1_weight_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='weight percent',variable=ingredient1_weight_percent_var,offvalue=False,onvalue=True)
            ingredient1_weight_percent_checkbutton.grid(row=11,column=6)
            ingredient1_volume_percent_var = tk.BooleanVar(value=False)
            ingredient1_volume_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='volume percent',variable=ingredient1_volume_percent_var,offvalue=False,onvalue=True)
            ingredient1_volume_percent_checkbutton.grid(row=11,column=7)

            ttk.Label(optimization_screen_frame,text='Ingredient 2:').grid(row=12,column=0)
            ingredient2_var = tk.StringVar()
            ingredient2_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient2_var)
            ingredient2_entry.grid(row=12,column=1)
            ingredient2_start_var = tk.StringVar()
            ingredient2_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient2_start_var)
            ingredient2_start_entry.grid(row=12,column=2)
            ingredient2_stop_var = tk.StringVar()
            ingredient2_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient2_stop_var)
            ingredient2_stop_entry.grid(row=12,column=3)
            ingredient2_pH_start_var = tk.StringVar()
            ingredient2_pH_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient2_pH_start_var)
            ingredient2_pH_start_entry.grid(row=12,column=4)
            ingredient2_pH_stop_var = tk.StringVar()
            ingredient2_pH_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient2_pH_stop_var)
            ingredient2_pH_stop_entry.grid(row=12,column=5)
            ingredient2_weight_percent_var = tk.BooleanVar(value=False)
            ingredient2_weight_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='weight percent',variable=ingredient2_weight_percent_var,offvalue=False,onvalue=True)
            ingredient2_weight_percent_checkbutton.grid(row=12,column=6)
            ingredient2_volume_percent_var = tk.BooleanVar(value=False)
            ingredient2_volume_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='volume percent',variable=ingredient2_volume_percent_var,offvalue=False,onvalue=True)
            ingredient2_volume_percent_checkbutton.grid(row=12,column=7)

            ttk.Label(optimization_screen_frame,text='Ingredient 3:').grid(row=13,column=0)
            ingredient3_var = tk.StringVar()
            ingredient3_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient3_var)
            ingredient3_entry.grid(row=13,column=1)
            ingredient3_start_var = tk.StringVar()
            ingredient3_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient3_start_var)
            ingredient3_start_entry.grid(row=13,column=2)
            ingredient3_stop_var = tk.StringVar()
            ingredient3_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient3_stop_var)
            ingredient3_stop_entry.grid(row=13,column=3)
            ingredient3_pH_start_var = tk.StringVar()
            ingredient3_pH_start_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient3_pH_start_var)
            ingredient3_pH_start_entry.grid(row=13,column=4)
            ingredient3_pH_stop_var = tk.StringVar()
            ingredient3_pH_stop_entry = tk.Entry(optimization_screen_frame,textvariable=ingredient3_pH_stop_var)
            ingredient3_pH_stop_entry.grid(row=13,column=5)
            ingredient3_weight_percent_var = tk.BooleanVar(value=False)
            ingredient3_weight_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='weight percent',variable=ingredient3_weight_percent_var,offvalue=False,onvalue=True)
            ingredient3_weight_percent_checkbutton.grid(row=13,column=6)
            ingredient3_volume_percent_var = tk.BooleanVar(value=False)
            ingredient3_volume_percent_checkbutton = ttk.Checkbutton(optimization_screen_frame,text='volume percent',variable=ingredient3_volume_percent_var,offvalue=False,onvalue=True)
            ingredient3_volume_percent_checkbutton.grid(row=13,column=7)
            
            conditions = ['' for _ in range(96)]
            conditions_var = tk.StringVar(value=conditions)
            conditions_listbox = tk.Listbox(optimization_screen_frame,listvariable=conditions_var,height=25,width=150)
            conditions_listbox.grid(row=14,column=0,columnspan=3)

            edited_condition = tk.StringVar()
            condition_entry = tk.Entry(optimization_screen_frame, textvariable=edited_condition, width=150)
            condition_entry.grid(row=15, column=0, columnspan=3)

            self.index = None
            def select_condition(event):
                selection = conditions_listbox.curselection()
                if selection:
                    edited_condition.set(f'{conditions[selection[0]+1]}')
                    self.index = selection[0]

            conditions_listbox.bind('<<ListboxSelect>>',select_condition)

            def overwrite():
                if self.index:
                    text = edited_condition.get()
                    conditions[self.index-1] = text
                    conditions_listbox.delete(self.index)
                    conditions_listbox.insert(self.index, f'{self.index+1} {text}')
                    
            tk.Button(optimization_screen_frame,text='overwrite',command=overwrite).grid(row=15,column=3)

            tk.Button(optimization_screen_frame,text='Add selection to custom screen',command=lambda:save_condition_settings(ingredient0_var.get(),ingredient0_start_var.get(),ingredient0_stop_var.get(),ingredient0_weight_percent_var.get(),ingredient0_pH_start_var.get(),ingredient0_pH_stop_var.get(),ingredient0_volume_percent_var.get(),ingredient1_var.get(),ingredient1_start_var.get(),ingredient1_stop_var.get(),ingredient1_weight_percent_var.get(),ingredient1_pH_start_var.get(),ingredient1_pH_stop_var.get(),ingredient1_volume_percent_var.get(),ingredient2_var.get(),ingredient2_start_var.get(),ingredient2_stop_var.get(),ingredient2_weight_percent_var.get(),ingredient2_pH_start_var.get(),ingredient2_pH_stop_var.get(),ingredient2_volume_percent_var.get(),ingredient3_var.get(),ingredient3_start_var.get(),ingredient3_stop_var.get(),ingredient3_weight_percent_var.get(),ingredient3_pH_start_var.get(),ingredient3_pH_stop_var.get(),ingredient3_volume_percent_var.get(),steps=int(steps_var.get()))).grid(row=50,column=0)

            tk.Button(optimization_screen_frame,text='Finish custom screen',command=lambda:save_screen()).grid(row=51,column=0)

            def save_condition_settings(ingredient0,ingredient0_start,ingredient0_stop,ingredient0_weight_percent,ingredient0_pH_start,ingredient0_pH_stop,ingredient0_volume_percent,
                                        ingredient1,ingredient1_start,ingredient1_stop,ingredient1_weight_percent,ingredient1_pH_start,ingredient1_pH_stop,ingredient1_volume_percent,
                                        ingredient2,ingredient2_start,ingredient2_stop,ingredient2_weight_percent,ingredient2_pH_start,ingredient2_pH_stop,ingredient2_volume_percent,
                                        ingredient3,ingredient3_start,ingredient3_stop,ingredient3_weight_percent,ingredient3_pH_start,ingredient3_pH_stop,ingredient3_volume_percent,
                                        steps=1):
                condition_instructions = [[ingredient0,ingredient0_start,ingredient0_stop,ingredient0_weight_percent,ingredient0_pH_start,ingredient0_pH_stop,ingredient0_volume_percent],[ingredient1,ingredient1_start,ingredient1_stop,ingredient1_weight_percent,ingredient1_pH_start,ingredient1_pH_stop,ingredient1_volume_percent],[ingredient2,ingredient2_start,ingredient2_stop,ingredient2_weight_percent,ingredient2_pH_start,ingredient2_pH_stop,ingredient2_volume_percent],[ingredient3,ingredient3_start,ingredient3_stop,ingredient3_weight_percent,ingredient3_pH_start,ingredient3_pH_stop,ingredient3_volume_percent]]
                current_condition_number = len(self.optimization_conditions.keys())
                for condition_number in range(current_condition_number,current_condition_number+steps):
                    if condition_number<=95:
                        self.optimization_conditions[condition_number] = f'{condition_number+1} '
                        for condition in condition_instructions:
                            if '' not in condition[0:4]:
                                new_ingredient_id = condition[0]
                                new_condition_start = float(condition[1])
                                new_condition_stop = float(condition[2])
                                new_condition_step = (new_condition_stop-new_condition_start)/(steps-1)
                                new_condition_concentration = condition_number*new_condition_step+new_condition_start
                                self.optimization_conditions[condition_number] = self.optimization_conditions[condition_number]+f'{round(new_condition_concentration,2)}'
                                if condition[3]:
                                    self.optimization_conditions[condition_number] = self.optimization_conditions[condition_number]+f' % w/v {new_ingredient_id} '
                                elif condition[6]:
                                    self.optimization_conditions[condition_number] = self.optimization_conditions[condition_number]+f' % v/v {new_ingredient_id} '                            
                                else:
                                    self.optimization_conditions[condition_number] = self.optimization_conditions[condition_number]+f' M {new_ingredient_id} '
                                if '' not in condition[4:6]:
                                    new_condition_pH = round(condition_number*(float(condition[5])-float(condition[4]))/(steps-1)+float(condition[4]),2)
                                    self.optimization_conditions[condition_number] = self.optimization_conditions[condition_number]+f' pH {new_condition_pH}, '
                    else:
                        messagebox.showerror(title="custom Conditions Full",message="There is no more room to add conditions to this screen.")

                for condition in range(len(self.optimization_conditions)):
                    conditions_listbox.delete(condition)
                    conditions_listbox.insert(condition, self.optimization_conditions[condition])

            def save_screen():
                crystal_screens = get_crystal_screens()
                crystal_screen_id = crystal_screens[self.crystal_screen_name]
                conn = connect_to_db()
                cur = conn.cursor()
                for i, condition in enumerate(self.optimization_conditions.values(),start=1):
                    cur.execute(
                        "INSERT INTO conditions (crystal_screen_id,condition_number,condition) VALUES (?, ?, ?)",
                        (crystal_screen_id, i, condition)
                    )
                conn.commit()
                conn.close()
                self.startup()

    def Crystal_Transfer(self):
        self.clear_widgets()
        self.add_menu()
        crystal_transfer_frame = ttk.Frame(self.root,padding="3 3 12 12")
        self.current_frame = crystal_transfer_frame
        crystal_transfer_frame.grid(column=0,row=0,sticky='nw')

        runs = get_runs()
        if len(runs)==0:
            messagebox.showerror(title="No Runs!",message="You can't transfer; you don't have any runs.")
            self.startup()
            return
        ttk.Label(crystal_transfer_frame,text=f"Current run: {max(runs)}").grid(column=1,row=1)
        ttk.Button(crystal_transfer_frame,text="Transfer crystals in current run",command=lambda:transfer(max(runs),method='current')).grid(column=1,row=2)
        ttk.Button(crystal_transfer_frame,text="Transfer crystals from an old run to the current run",command=lambda:transfer(method='old')).grid(column=1,row=4)

        def transfer(run=None,method=None):
            self.clear_widgets()
            self.add_menu()
            transfer_frame = ttk.Frame(self.root,padding="3 3 12 12")
            self.current_frame = transfer_frame
            transfer_frame.grid(column=0,row=0,sticky='nw')

            if run==None:
                run = tk.StringVar()
                ttk.Label(transfer_frame, text="Select the old run to transfer crystals from:").grid(column=1,row=1)
                ttk.Combobox(transfer_frame,values=runs,textvariable=run).grid(column=1,row=2)
                tk.Button(transfer_frame,text="Transfer from selected tray",command=lambda:transfer(int(run.get()),method='old')).grid(column=1,row=3)
                return

            wi = 330

            vials_and_ports = get_vials_and_details(run)
            transfer_tree=None

            if method=="current":
                transfer_tree = ttk.Treeview(
                    transfer_frame,
                    columns=("vial", "port", "protein", "conditions", "notes"),
                    show="headings",
                    height=25
                )

                transfer_tree.heading("vial", text="Vial")
                transfer_tree.heading("port", text="Port")
                transfer_tree.heading("protein", text="Protein")
                transfer_tree.heading("conditions", text="Conditions")
                transfer_tree.heading("notes", text="Notes")
                transfer_tree.column("vial", width=wi)
                transfer_tree.column("port", width=wi)
                transfer_tree.column("protein", width=wi)
                transfer_tree.column("conditions", width=wi)
                transfer_tree.column("notes", width=wi)

            elif method=="old":
                transfer_tree = ttk.Treeview(
                    transfer_frame,
                    columns=("vial", "transfer_b", "protein", "conditions", "notes"),
                    show="headings",
                    height=25
                )
                transfer_tree.heading("vial", text="Vial")
                transfer_tree.heading("transfer_b",text="Transfer to current run?")
                transfer_tree.heading("protein", text="Protein")
                transfer_tree.heading("conditions", text="Conditions")
                transfer_tree.heading("notes", text="Notes")
                transfer_tree.column("vial", width=wi)
                transfer_tree.column("transfer_b",width=wi)
                transfer_tree.column("protein", width=wi)
                transfer_tree.column("conditions", width=wi)
                transfer_tree.column("notes", width=wi)

            transfer_tree.grid(row=0, column=0)

            for vial, details in vials_and_ports.items():
                if method=="current":
                    transfer_tree.insert("","end",values=(vial,details[0],details[1],details[2],details[3]))
                elif method=="old":
                    transfer_tree.insert("","end",values=(vial,"False",details[1],details[2],details[3]))
            scrollbar = ttk.Scrollbar(transfer_frame, orient="vertical",command=transfer_tree.yview)
            transfer_tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.grid(row=0,column=1,sticky="ns")

            def on_select(event):
                if method=="current":
                    selected = transfer_tree.selection()
                    if not selected:
                        return
                    item = selected[0]
                    #values = transfer_tree.item(item, "values")

                    x, y, width, height = transfer_tree.bbox(item, "#2")

                    value = transfer_tree.set(item, "port")

                    entry = ttk.Entry(transfer_tree)
                    entry.place(x=x, y=y, width=width, height=height)
                    entry.insert(0, value)
                    entry.focus()
                    def save_edit(event=None):
                        transfer_tree.set(item, "port", entry.get())
                        entry.destroy()
                    entry.bind("<Return>", save_edit)
                    entry.bind("<FocusOut>", save_edit)
                elif method=="old":
                    selected = transfer_tree.selection()
                    if not selected:
                        return
                    item = selected[0]
                    x, y, width, height = transfer_tree.bbox(item, "#2")
                    transfer_b = tk.BooleanVar(value=False)
                    entry = ttk.Checkbutton(transfer_tree,variable=transfer_b,offvalue=False,onvalue=True)
                    entry.place(x=x, y=y, width=width, height=height)
                    entry.focus()
                    def save_edit(event=None):
                        if transfer_b.get()==0:
                            transfer_tree.set(item, "transfer_b", "False")
                        elif transfer_b.get()==1:
                            transfer_tree.set(item, "transfer_b", "True")
                        entry.destroy()
                    entry.bind("<Return>", save_edit)
                    entry.bind("<FocusOut>", save_edit)
            transfer_tree.bind("<<TreeviewSelect>>", on_select)

            if method=='current':
                ttk.Button(transfer_frame,text="Save, Print Run Sheet, and Return to Startup Menu",command=lambda:save()).grid(column=0,row=2)
            elif method=='old':
                ttk.Button(transfer_frame,text="Move crystals to current run",command=lambda:move()).grid(column=0,row=2)

            def save():
                updates = []
                for item in transfer_tree.get_children():
                    vial, port, protein, conditions, notes = transfer_tree.item(item, "values")
                    updates.append((port, run, int(vial)))
                print(f'updates: {updates}')
                conn = connect_to_db()
                cur = conn.cursor()
                cur.executemany("""UPDATE crystals SET port=? WHERE run=? AND vial=?""",updates)
                conn.commit()
                conn.close()
                self.run_sheet(run)
                self.startup()

            def move():
                current_run_vials_and_ports = get_vials_and_details(max(runs))
                next_vial = 0
                for vial in current_run_vials_and_ports.keys():
                    if int(vial)>next_vial:
                        next_vial = int(vial)+1
                updates = []
                for item in transfer_tree.get_children():
                    vial, transfer_b, protein, conditions, notes = transfer_tree.item(item, "values")
                    new_note = f"Old run and vial: ({run}, {vial})"
                    if transfer_b=="True":
                        updates.append((next_vial, max(runs), new_note, run, int(vial)))
                        next_vial+=1
                print(f'updates: {updates}')
                conn = connect_to_db()
                cur = conn.cursor()
                cur.executemany("""UPDATE crystals SET vial = ?, run=?, notes = ? || notes WHERE run=? AND vial=?""",updates)
                conn.commit()
                conn.close()
                self.run_sheet(run)
                self.startup()

    def run_sheet(self,run):
        """This creates and prints an Excel sheet depicting each of the crystals in the given run."""
        conn = connect_to_db()

        df = pd.read_sql_query("""SELECT
        c.vial AS Vial,
        t.protein AS Protein,
        c.conditions AS Conditions,
        c.shape AS Shape,
        c.minor_axis AS Minor_axis,
        c.major_axis AS Major_axis,
        c.harvester AS Harvester,
        c.notes AS Notes,
        c.run AS Run,
        t.id AS tray_id,
        c.tray_id AS Crystal_tray_id,
        t.date_set AS Date_set,
        c.date_snapped AS Date_snapped,
        c.port as Port
        FROM crystals c JOIN crystal_trays t ON t.id = c.tray_id""", conn)
        conn.close()

        df = df[df["Run"] == run]

        with pd.ExcelWriter(run_sheet, engine="openpyxl",mode="w") as writer:
            if df.empty:
                pd.DataFrame({"info": ["No data available"]}).to_excel(
                    writer,
                    sheet_name="Empty",
                    index=False
                )
            else:
                """
                df['Picture'] = df['Picture'].apply(
                    lambda x: f'=HYPERLINK("{os.path.join(crystal_pictures,x)}", "Open Image")' if pd.notnull(x) else ""
                )
                """
                df.insert(7,"Date Set/Date harvested",df["Date_set"].astype(str) + "/" + df["Date_snapped"].astype(str))
                base_name = f"Run {run}"
                sheet_name = base_name[:31]
                df = df.drop(columns=["Run","tray_id","Crystal_tray_id","Date_set","Date_snapped"])
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=0,
                    index=False
                )
